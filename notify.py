from __future__ import annotations

import argparse
import asyncio
import csv
import json
import sys
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import aiohttp
from openpyxl import load_workbook


DEFAULT_CONFIG_PATH = Path(
    r"C:\ProgramData\NapCatQQ Desktop\runtime\NapCatQQ\config\onebot11_3496930386.json"
)
DEFAULT_OUTPUT_ROOT = Path(r"C:\Users\Theta\napcat_notify\runs")
DEFAULT_MESSAGE_TEMPLATE = (
    "{name}同学你好，我是乒协机器人，这边发现你在正赛资格名单中，"
    "但单打项目报名表中没有你的信息。请你确认是否参加成电杯单打项目，"
    "如果参加，请填写群里的单打报名链接。"
)
RESULT_FIELDS = [
    "qq",
    "name",
    "college",
    "mode",
    "status",
    "message_id",
    "error",
    "sent_at",
    "message_preview",
]


@dataclass
class Recipient:
    qq: str
    name: str
    college: str
    message: str | None = None


@dataclass
class ResultRow:
    qq: str
    name: str
    college: str
    mode: str
    status: str
    message_id: str
    error: str
    sent_at: str
    message_preview: str


DEFAULT_RECIPIENTS = [
    Recipient("3331411486", "张睿康", "电子科学与工程学院"),
    Recipient("419937075", "王毓远", "电子科学与工程学院"),
    Recipient("2428336731", "廖鸣宇", "格拉斯哥学院（清水河）"),
    Recipient("3179989582", "郭泰佑", "航空航天学院"),
    Recipient("1656925659", "李泊宇", "通信抗干扰全国重点实验室"),
    Recipient("760837562", "张艺宝", "外国语学院"),
    Recipient("489842854", "陈子信", "物理学院"),
    Recipient("1260346350", "李昊越", "信息与软件工程学院（示范性软件学院）"),
    Recipient("836302957", "周尚楠", "信息与通信工程学院"),
    Recipient("3562814086", "王言立", "自动化工程学院"),
    Recipient("1194073522", "徐宇轩", "自动化工程学院"),
    Recipient("1169370631", "周舒琪", "材料与能源学院"),
    Recipient("2689004746", "王艺锦", "光电科学与工程学院"),
    Recipient("2420729267", "易若涵", "计算机科学与工程学院（网络空间安全学院）"),
    Recipient("3466541652", "柯馨怡", "信息与通信工程学院"),
    Recipient("2217499841", "席子婷", "信息与通信工程学院"),
    Recipient("3831402618", "周丝怡", "信息与通信工程学院"),
    Recipient("1841974993", "肖珍佳", "研究生院"),
    Recipient("1052051115", "张浩月", "医学院"),
    Recipient("1204890883", "周杨茜", "英才实验学院（未来技术学院）"),
    Recipient("2632385062", "黄子畔", "英才实验学院（未来技术学院）"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="NapCat 批量私聊通知脚本")
    parser.add_argument("--input", type=Path, help="名单文件路径，支持 .csv / .xlsx")
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="NapCat OneBot11 配置文件路径",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help="运行结果输出目录根路径",
    )
    parser.add_argument("--send", action="store_true", help="真正发送消息")
    parser.add_argument("--limit", type=int, help="只处理前 N 条记录")
    parser.add_argument("--delay", type=float, default=2.0, help="每条消息间隔秒数")
    parser.add_argument(
        "--group-id",
        type=int,
        help="启用群临时会话发送时使用的群号；提供后将以群临时会话方式发私聊",
    )
    return parser.parse_args()


def load_ws_server_config(config_path: Path) -> dict[str, Any]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    servers = config.get("network", {}).get("websocketServers", [])
    enabled = [item for item in servers if item.get("enable")]
    if not enabled:
        raise RuntimeError(f"未在 {config_path} 中找到已启用的 websocketServers 配置")
    server = enabled[0]
    for key in ("host", "port", "token"):
        if not server.get(key):
            raise RuntimeError(f"WebSocket 服务端配置缺少字段: {key}")
    return server


def load_recipients(path: Path | None) -> list[Recipient]:
    if path is None:
        return list(DEFAULT_RECIPIENTS)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_recipients_from_csv(path)
    if suffix == ".xlsx":
        return load_recipients_from_xlsx(path)
    raise RuntimeError(f"不支持的文件类型: {path.suffix}")


def load_recipients_from_csv(path: Path) -> list[Recipient]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        return parse_recipient_rows(reader, source=str(path))


def load_recipients_from_xlsx(path: Path) -> list[Recipient]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    dict_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(value is not None and str(value).strip() for value in row):
            continue
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            item[header] = row[idx] if idx < len(row) else None
        dict_rows.append(item)
    return parse_recipient_rows(dict_rows, source=str(path))


def parse_recipient_rows(rows: list[dict[str, Any]] | Any, source: str) -> list[Recipient]:
    recipients: list[Recipient] = []
    for index, row in enumerate(rows, start=2):
        qq_raw = row.get("QQ号")
        name_raw = row.get("姓名")
        college_raw = row.get("学院")
        message_raw = row.get("消息")
        qq = normalize_qq(qq_raw)
        name = normalize_text(name_raw)
        college = normalize_text(college_raw)
        message = normalize_text(message_raw)
        if not qq or not name or not college:
            raise RuntimeError(f"{source} 第 {index} 行缺少必要列值")
        recipients.append(Recipient(qq=qq, name=name, college=college, message=message or None))
    if not recipients:
        raise RuntimeError(f"{source} 未读取到任何收件人")
    return recipients


def normalize_qq(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        raise RuntimeError(f"QQ号不是纯数字: {value!r}")
    return digits


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_message(recipient: Recipient) -> str:
    template = recipient.message or DEFAULT_MESSAGE_TEMPLATE
    return template.format(name=recipient.name)


def truncate_preview(message: str, width: int = 80) -> str:
    if len(message) <= width:
        return message
    return message[: width - 3] + "..."


def make_run_dir(output_root: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = output_root / stamp
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def write_results(run_dir: Path, rows: list[ResultRow]) -> None:
    csv_path = run_dir / "results.csv"
    jsonl_path = run_dir / "results.jsonl"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=RESULT_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))
    with jsonl_path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(asdict(row), ensure_ascii=False) + "\n")


class NapCatWsClient:
    def __init__(self, host: str, port: int, token: str) -> None:
        self.url = f"ws://{host}:{port}"
        self.headers = {"Authorization": f"Bearer {token}"}
        self.session: aiohttp.ClientSession | None = None
        self.ws: aiohttp.ClientWebSocketResponse | None = None

    async def __aenter__(self) -> "NapCatWsClient":
        timeout = aiohttp.ClientTimeout(total=20)
        self.session = aiohttp.ClientSession(timeout=timeout)
        self.ws = await self.session.ws_connect(self.url, headers=self.headers)
        return self

    async def __aexit__(self, exc_type, exc, tb) -> None:
        if self.ws is not None:
            await self.ws.close()
        if self.session is not None:
            await self.session.close()

    async def request(self, action: str, params: dict[str, Any] | None = None, timeout: float = 15.0) -> dict[str, Any]:
        if self.ws is None:
            raise RuntimeError("WebSocket 尚未连接")
        echo = uuid.uuid4().hex
        payload = {"action": action, "params": params or {}, "echo": echo}
        await self.ws.send_json(payload)
        deadline = asyncio.get_running_loop().time() + timeout
        while True:
            remaining = deadline - asyncio.get_running_loop().time()
            if remaining <= 0:
                raise TimeoutError(f"{action} 响应超时")
            msg = await self.ws.receive(timeout=remaining)
            if msg.type == aiohttp.WSMsgType.TEXT:
                data = json.loads(msg.data)
                if data.get("echo") == echo:
                    return data
                continue
            if msg.type in (aiohttp.WSMsgType.CLOSED, aiohttp.WSMsgType.CLOSE, aiohttp.WSMsgType.ERROR):
                raise RuntimeError(f"WebSocket 已关闭: {msg.type}")


def response_error_text(response: dict[str, Any]) -> str:
    message = response.get("message")
    retcode = response.get("retcode")
    status = response.get("status")
    if message or retcode or status:
        return f"status={status!r}, retcode={retcode!r}, message={message!r}"
    return json.dumps(response, ensure_ascii=False)


async def verify_group_member(
    client: NapCatWsClient, group_id: int, user_id: int
) -> tuple[bool, str]:
    response = await client.request(
        "get_group_member_info",
        {"group_id": group_id, "user_id": user_id, "no_cache": True},
    )
    if response.get("status") == "ok" and response.get("retcode") == 0:
        return True, ""
    return False, response_error_text(response)


async def run(args: argparse.Namespace) -> int:
    recipients = load_recipients(args.input)
    if args.limit is not None:
        recipients = recipients[: args.limit]
    if not recipients:
        raise RuntimeError("没有可处理的收件人")

    ws_config = load_ws_server_config(args.config)
    run_dir = make_run_dir(args.output_root)
    mode = "send" if args.send else "dry_run"

    print(f"运行模式: {mode}")
    print(f"收件人数: {len(recipients)}")
    print(f"运行目录: {run_dir}")
    print(f"WebSocket: ws://{ws_config['host']}:{ws_config['port']}")
    if args.group_id:
        print(f"发送方式: 群临时会话 (group_id={args.group_id})")
    else:
        print("发送方式: 普通私聊")
    print()
    for index, recipient in enumerate(recipients, start=1):
        print(f"[{index:02d}] {recipient.name} ({recipient.qq}) - {recipient.college}")
        print(f"     {truncate_preview(build_message(recipient), 120)}")

    preflight_error = ""
    login_info: dict[str, Any] | None = None
    try:
        async with NapCatWsClient(
            host=ws_config["host"],
            port=int(ws_config["port"]),
            token=str(ws_config["token"]),
        ) as client:
            login_response = await client.request("get_login_info")
            if login_response.get("status") != "ok" or login_response.get("retcode") != 0:
                raise RuntimeError("get_login_info 失败: " + response_error_text(login_response))
            login_info = login_response.get("data") or {}
            print()
            print(f"连接预检成功，当前 bot: {login_info.get('user_id')} / {login_info.get('nickname')}")

            if not args.send:
                rows = [
                    ResultRow(
                        qq=recipient.qq,
                        name=recipient.name,
                        college=recipient.college,
                        mode=mode,
                        status="dry_run",
                        message_id="",
                        error="",
                        sent_at="",
                        message_preview=truncate_preview(build_message(recipient), 120),
                    )
                    for recipient in recipients
                ]
                write_results(run_dir, rows)
                return 0

            rows: list[ResultRow] = []
            for index, recipient in enumerate(recipients, start=1):
                message = build_message(recipient)
                sent_at = datetime.now().isoformat(timespec="seconds")
                try:
                    params: dict[str, Any] = {
                        "user_id": int(recipient.qq),
                        "message": message,
                    }
                    if args.group_id:
                        is_member, member_error = await verify_group_member(
                            client, args.group_id, int(recipient.qq)
                        )
                        if not is_member:
                            rows.append(
                                ResultRow(
                                    qq=recipient.qq,
                                    name=recipient.name,
                                    college=recipient.college,
                                    mode=mode,
                                    status="api_failed",
                                    message_id="",
                                    error=f"group_member_check_failed: {member_error}",
                                    sent_at=sent_at,
                                    message_preview=truncate_preview(message, 120),
                                )
                            )
                            print(
                                f"[{index:02d}] 群成员校验失败: {recipient.name} ({recipient.qq}) -> {member_error}"
                            )
                            if index < len(recipients):
                                await asyncio.sleep(args.delay)
                            continue
                        params["group_id"] = int(args.group_id)
                    response = await client.request(
                        "send_private_msg",
                        params,
                    )
                    if response.get("status") == "ok" and response.get("retcode") == 0:
                        data = response.get("data") or {}
                        rows.append(
                            ResultRow(
                                qq=recipient.qq,
                                name=recipient.name,
                                college=recipient.college,
                                mode=mode,
                                status="sent",
                                message_id=str(data.get("message_id", "")),
                                error="",
                                sent_at=sent_at,
                                message_preview=truncate_preview(message, 120),
                            )
                        )
                        print(f"[{index:02d}] 已发送: {recipient.name} ({recipient.qq})")
                    else:
                        error = response_error_text(response)
                        rows.append(
                            ResultRow(
                                qq=recipient.qq,
                                name=recipient.name,
                                college=recipient.college,
                                mode=mode,
                                status="api_failed",
                                message_id="",
                                error=error,
                                sent_at=sent_at,
                                message_preview=truncate_preview(message, 120),
                            )
                        )
                        print(f"[{index:02d}] 发送失败: {recipient.name} ({recipient.qq}) -> {error}")
                except Exception as exc:
                    rows.append(
                        ResultRow(
                            qq=recipient.qq,
                            name=recipient.name,
                            college=recipient.college,
                            mode=mode,
                            status="transport_failed",
                            message_id="",
                            error=str(exc),
                            sent_at=sent_at,
                            message_preview=truncate_preview(message, 120),
                        )
                    )
                    print(f"[{index:02d}] 连接失败: {recipient.name} ({recipient.qq}) -> {exc}")
                if index < len(recipients):
                    await asyncio.sleep(args.delay)
            write_results(run_dir, rows)
            return 0
    except Exception as exc:
        preflight_error = str(exc)
        print()
        print(f"连接预检失败: {preflight_error}", file=sys.stderr)

    rows = []
    for recipient in recipients:
        rows.append(
            ResultRow(
                qq=recipient.qq,
                name=recipient.name,
                college=recipient.college,
                mode=mode,
                status="transport_failed" if args.send else "dry_run",
                message_id="",
                error=preflight_error if args.send else "",
                sent_at="",
                message_preview=truncate_preview(build_message(recipient), 120),
            )
        )
    write_results(run_dir, rows)
    return 1 if args.send else 0


def main() -> int:
    args = parse_args()
    try:
        return asyncio.run(run(args))
    except KeyboardInterrupt:
        print("用户中断")
        return 130
    except Exception as exc:
        print(f"执行失败: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
