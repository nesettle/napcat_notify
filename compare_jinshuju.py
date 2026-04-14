from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from tempfile import mkdtemp
from typing import Any

from openpyxl import load_workbook
from playwright.sync_api import BrowserContext, Error, Locator, Page, sync_playwright


DEFAULT_OUTPUT_ROOT = Path(r"C:\Users\Theta\napcat_notify\runs")
DEFAULT_PROFILE_DIR = Path(r"C:\Users\Theta\napcat_notify\browser_state\jinshuju_profile")
DEFAULT_DOWNLOAD_FORMAT = "xlsx"
DEFAULT_HOME_URL = "https://jinshuju.net/home"


@dataclass
class QualificationRecord:
    source: str
    sheet: str
    row_number: int
    name: str
    college: str
    qq: str
    normalized_name: str
    normalized_college: str
    normalized_qq: str


@dataclass
class FormRecord:
    source: str
    row_number: int
    created_at: str
    updated_at: str
    serial_number: str
    name: str
    college: str
    qq: str
    normalized_name: str
    normalized_college: str
    normalized_qq: str


@dataclass
class DuplicateRow:
    source: str
    key_type: str
    match_key: str
    name: str
    college: str
    qq: str
    details: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从金数据后台导出报名表并与资格名单比对")
    parser.add_argument("--form-title", help="金数据首页中的表单标题")
    parser.add_argument("--entries-url", help="金数据数据页直达链接，例如 https://jinshuju.net/forms/<token>/entries")
    parser.add_argument(
        "--qualification-file",
        required=True,
        type=Path,
        help="本地资格名单 Excel 文件路径",
    )
    parser.add_argument("--name-field-label", default="姓名", help="导出表中的姓名列标题")
    parser.add_argument("--college-field-label", default="学院", help="导出表中的学院列标题")
    parser.add_argument("--qq-field-label", default="QQ号", help="导出表中的 QQ 列标题")
    parser.add_argument(
        "--created-after",
        help="只统计该时间之后创建的报名记录，支持 YYYY-MM-DD 或 ISO 8601",
    )
    parser.add_argument(
        "--download-format",
        choices=("xlsx", "csv"),
        default=DEFAULT_DOWNLOAD_FORMAT,
        help="从金数据导出的文件格式",
    )
    parser.add_argument(
        "--profile-dir",
        type=Path,
        default=DEFAULT_PROFILE_DIR,
        help="浏览器持久化登录目录",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        default=DEFAULT_OUTPUT_ROOT,
        help="结果输出目录根路径",
    )
    parser.add_argument(
        "--home-url",
        default=DEFAULT_HOME_URL,
        help="金数据后台首页地址",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="无头模式运行；首次登录建议不要使用",
    )
    args = parser.parse_args()
    if not args.form_title and not args.entries_url:
        parser.error("--form-title 和 --entries-url 至少需要提供一个")
    return args


def parse_user_datetime(value: str) -> datetime:
    text = value.strip()
    if not text:
        raise ValueError("时间参数不能为空")
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return datetime.fromisoformat(text).replace(tzinfo=timezone.utc)
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def parse_timestamp(value: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    for candidate in (text, text.replace("/", "-")):
        try:
            dt = datetime.fromisoformat(candidate)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            continue
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def collapse_spaces(value: str) -> str:
    return " ".join(value.split())


def normalize_name(value: Any) -> str:
    text = collapse_spaces(str(value or "").strip())
    while True:
        updated = re.sub(r"\s*[（(][^()（）]*[)）]\s*$", "", text)
        if updated == text:
            break
        text = updated.strip()
    return text


def normalize_college(value: Any) -> str:
    text = collapse_spaces(str(value or "").strip())
    return text.replace("（", "(").replace("）", ")")


def normalize_qq(value: Any) -> str:
    text = str(value or "").strip()
    return "".join(ch for ch in text if ch.isdigit())


def detect_qualification_columns(headers: list[str]) -> tuple[int, int, int]:
    name_idx = -1
    college_idx = -1
    qq_idx = -1
    for idx, header in enumerate(headers):
        normalized = collapse_spaces(header).replace("（", "(").replace("）", ")")
        if name_idx < 0 and normalized == "姓名":
            name_idx = idx
        if college_idx < 0 and "学院" in normalized:
            college_idx = idx
        if qq_idx < 0 and normalized == "QQ号":
            qq_idx = idx
    if name_idx < 0 or college_idx < 0 or qq_idx < 0:
        raise RuntimeError("资格名单表头解析失败，必须包含 姓名 / QQ号 / 包含“学院”的列")
    return name_idx, college_idx, qq_idx


def load_qualification_records(path: Path) -> list[QualificationRecord]:
    workbook = load_workbook(path, data_only=True)
    records: list[QualificationRecord] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [collapse_spaces(str(cell or "").strip()) for cell in rows[0]]
        if not any(headers):
            continue
        name_idx, college_idx, qq_idx = detect_qualification_columns(headers)
        for row_number, row in enumerate(rows[1:], start=2):
            values = list(row)
            name = str(values[name_idx] or "").strip() if name_idx < len(values) else ""
            college = str(values[college_idx] or "").strip() if college_idx < len(values) else ""
            qq = str(values[qq_idx] or "").strip() if qq_idx < len(values) else ""
            if not any((name, college, qq)):
                continue
            records.append(
                QualificationRecord(
                    source=str(path),
                    sheet=sheet.title,
                    row_number=row_number,
                    name=name,
                    college=college,
                    qq=qq,
                    normalized_name=normalize_name(name),
                    normalized_college=normalize_college(college),
                    normalized_qq=normalize_qq(qq),
                )
            )
    if not records:
        raise RuntimeError(f"未从 {path} 读取到任何资格名单记录")
    return records


def read_csv_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [row for row in csv.reader(fh)]


def read_excel_rows(path: Path) -> list[list[str]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value).strip() for value in row])
    return rows


def detect_export_columns(headers: list[str], name_label: str, college_label: str, qq_label: str) -> dict[str, int]:
    mapping = {
        name_label: -1,
        college_label: -1,
        qq_label: -1,
        "序号": -1,
        "提交时间": -1,
        "更新时间": -1,
    }
    normalized_headers = [collapse_spaces(item).replace("（", "(").replace("）", ")") for item in headers]
    for idx, header in enumerate(normalized_headers):
        if mapping[name_label] < 0 and header == collapse_spaces(name_label):
            mapping[name_label] = idx
        if mapping[college_label] < 0 and header == collapse_spaces(college_label):
            mapping[college_label] = idx
        if mapping[qq_label] < 0 and header == collapse_spaces(qq_label):
            mapping[qq_label] = idx
        if mapping["序号"] < 0 and header == "序号":
            mapping["序号"] = idx
        if mapping["提交时间"] < 0 and header in {"提交时间", "创建时间"}:
            mapping["提交时间"] = idx
        if mapping["更新时间"] < 0 and header in {"更新时间", "最后更新时间"}:
            mapping["更新时间"] = idx
    missing = [label for label in (name_label, college_label, qq_label) if mapping[label] < 0]
    if missing:
        raise RuntimeError(f"导出表中缺少必要列: {', '.join(missing)}")
    return mapping


def load_form_records_from_export(
    path: Path,
    name_label: str,
    college_label: str,
    qq_label: str,
    created_after: datetime | None,
) -> list[FormRecord]:
    rows = read_excel_rows(path) if path.suffix.lower() == ".xlsx" else read_csv_rows(path)
    if not rows:
        raise RuntimeError(f"导出文件为空: {path}")
    headers = rows[0]
    column_map = detect_export_columns(headers, name_label, college_label, qq_label)
    results: list[FormRecord] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        name = row[column_map[name_label]] if column_map[name_label] < len(row) else ""
        college = row[column_map[college_label]] if column_map[college_label] < len(row) else ""
        qq = row[column_map[qq_label]] if column_map[qq_label] < len(row) else ""
        created_at = row[column_map["提交时间"]] if 0 <= column_map["提交时间"] < len(row) else ""
        updated_at = row[column_map["更新时间"]] if 0 <= column_map["更新时间"] < len(row) else ""
        serial_number = row[column_map["序号"]] if 0 <= column_map["序号"] < len(row) else ""
        created_dt = parse_timestamp(created_at)
        if created_after and created_dt and created_dt < created_after:
            continue
        record = FormRecord(
            source=str(path),
            row_number=row_number,
            created_at=created_at,
            updated_at=updated_at,
            serial_number=serial_number,
            name=name,
            college=college,
            qq=qq,
            normalized_name=normalize_name(name),
            normalized_college=normalize_college(college),
            normalized_qq=normalize_qq(qq),
        )
        if record.normalized_name or record.normalized_college or record.normalized_qq:
            results.append(record)
    if not results:
        raise RuntimeError(f"导出文件中没有可用报名记录: {path}")
    return results


def primary_key(normalized_qq: str, normalized_name: str, normalized_college: str) -> tuple[str, str]:
    if normalized_qq:
        return ("qq", normalized_qq)
    return ("name_college", f"{normalized_name}|{normalized_college}")


def dedupe_qualifications(records: list[QualificationRecord]) -> tuple[list[QualificationRecord], list[DuplicateRow]]:
    groups: dict[tuple[str, str], list[QualificationRecord]] = defaultdict(list)
    for record in records:
        groups[primary_key(record.normalized_qq, record.normalized_name, record.normalized_college)].append(record)
    deduped: list[QualificationRecord] = []
    duplicates: list[DuplicateRow] = []
    for (key_type, match_key), items in groups.items():
        deduped.append(items[0])
        if len(items) > 1:
            for item in items:
                duplicates.append(
                    DuplicateRow(
                        source=f"{item.source}#{item.sheet}:{item.row_number}",
                        key_type=key_type,
                        match_key=match_key,
                        name=item.name,
                        college=item.college,
                        qq=item.qq,
                        details="资格名单存在重复匹配键",
                    )
                )
    return deduped, duplicates


def sort_form_duplicates(items: list[FormRecord]) -> list[FormRecord]:
    def sort_key(item: FormRecord) -> tuple[datetime, int, int]:
        updated = parse_timestamp(item.updated_at) or parse_timestamp(item.created_at)
        if updated is None:
            updated = datetime.min.replace(tzinfo=timezone.utc)
        serial = int(item.serial_number) if item.serial_number.isdigit() else -1
        return (updated, serial, -item.row_number)

    return sorted(items, key=sort_key, reverse=True)


def dedupe_form_records(records: list[FormRecord]) -> tuple[list[FormRecord], list[DuplicateRow]]:
    groups: dict[tuple[str, str], list[FormRecord]] = defaultdict(list)
    for record in records:
        groups[primary_key(record.normalized_qq, record.normalized_name, record.normalized_college)].append(record)
    deduped: list[FormRecord] = []
    duplicates: list[DuplicateRow] = []
    for (key_type, match_key), items in groups.items():
        sorted_items = sort_form_duplicates(items)
        deduped.append(sorted_items[0])
        if len(sorted_items) > 1:
            for item in sorted_items:
                duplicates.append(
                    DuplicateRow(
                        source=f"{item.source}#{item.row_number}",
                        key_type=key_type,
                        match_key=match_key,
                        name=item.name,
                        college=item.college,
                        qq=item.qq,
                        details="金数据导出中存在重复报名，主比较仅保留最新一条",
                    )
                )
    return deduped, duplicates


def build_map_by_qq(records: list[Any]) -> dict[str, Any]:
    return {record.normalized_qq: record for record in records if record.normalized_qq}


def build_map_by_name_college(records: list[Any]) -> dict[tuple[str, str], Any]:
    return {
        (record.normalized_name, record.normalized_college): record
        for record in records
        if record.normalized_name and record.normalized_college
    }


def compare_records(
    qualifications: list[QualificationRecord],
    form_records: list[FormRecord],
) -> tuple[list[dict[str, str]], list[QualificationRecord], list[FormRecord]]:
    remaining_qual = qualifications[:]
    remaining_form = form_records[:]
    matched: list[dict[str, str]] = []

    qual_by_qq = build_map_by_qq(remaining_qual)
    form_by_qq = build_map_by_qq(remaining_form)
    matched_qual_ids: set[int] = set()
    matched_form_ids: set[int] = set()

    for qq, qual in qual_by_qq.items():
        form = form_by_qq.get(qq)
        if not form:
            continue
        matched_qual_ids.add(id(qual))
        matched_form_ids.add(id(form))
        matched.append(
            {
                "match_type": "qq",
                "qualification_name": qual.name,
                "qualification_college": qual.college,
                "qualification_qq": qual.qq,
                "qualification_source": f"{qual.sheet}:{qual.row_number}",
                "form_name": form.name,
                "form_college": form.college,
                "form_qq": form.qq,
                "form_source": f"{Path(form.source).name}:{form.row_number}",
                "form_created_at": form.created_at,
                "form_updated_at": form.updated_at,
            }
        )

    remaining_qual = [item for item in remaining_qual if id(item) not in matched_qual_ids]
    remaining_form = [item for item in remaining_form if id(item) not in matched_form_ids]

    form_name_map = build_map_by_name_college(remaining_form)
    fallback_form_ids: set[int] = set()
    fallback_qual_ids: set[int] = set()
    for qual in remaining_qual:
        key = (qual.normalized_name, qual.normalized_college)
        if not all(key):
            continue
        form = form_name_map.get(key)
        if not form or id(form) in fallback_form_ids:
            continue
        if qual.normalized_qq and form.normalized_qq:
            continue
        fallback_form_ids.add(id(form))
        fallback_qual_ids.add(id(qual))
        matched.append(
            {
                "match_type": "name_college_fallback",
                "qualification_name": qual.name,
                "qualification_college": qual.college,
                "qualification_qq": qual.qq,
                "qualification_source": f"{qual.sheet}:{qual.row_number}",
                "form_name": form.name,
                "form_college": form.college,
                "form_qq": form.qq,
                "form_source": f"{Path(form.source).name}:{form.row_number}",
                "form_created_at": form.created_at,
                "form_updated_at": form.updated_at,
            }
        )

    remaining_qual = [item for item in remaining_qual if id(item) not in fallback_qual_ids]
    remaining_form = [item for item in remaining_form if id(item) not in fallback_form_ids]
    return matched, remaining_qual, remaining_form


def ensure_run_dir(output_root: Path) -> Path:
    stamp = datetime.now().strftime("compare-%Y%m%d-%H%M%S")
    run_dir = output_root / stamp
    run_dir.mkdir(parents=True, exist_ok=False)
    return run_dir


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = list(rows[0].keys()) if rows else ["source"]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def qualification_to_csv_row(record: QualificationRecord) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{record.source}#{record.sheet}:{record.row_number}",
    }


def form_to_csv_row(record: FormRecord) -> dict[str, str]:
    return {
        "name": record.name,
        "college": record.college,
        "qq": record.qq,
        "source": f"{record.source}#{record.row_number}",
        "created_at": record.created_at,
        "updated_at": record.updated_at,
        "serial_number": record.serial_number,
    }


def duplicate_to_csv_row(record: DuplicateRow) -> dict[str, str]:
    return asdict(record)


def save_debug_artifacts(page: Page, run_dir: Path, name: str) -> None:
    html_path = run_dir / f"{name}.html"
    screenshot_path = run_dir / f"{name}.png"
    html_path.write_text(page.content(), encoding="utf-8")
    page.screenshot(path=str(screenshot_path), full_page=True)


def click_first_visible(locator: Locator, timeout_ms: int = 10000) -> bool:
    last_error: Exception | None = None
    try:
        count = locator.count()
    except Exception as exc:
        raise RuntimeError("无法统计可点击元素数量") from exc
    for idx in range(count):
        candidate = locator.nth(idx)
        try:
            if not candidate.is_visible():
                continue
            candidate.click(timeout=timeout_ms)
            return True
        except Exception as exc:
            last_error = exc
    if last_error:
        raise RuntimeError("找到候选元素但点击失败") from last_error
    return False


def has_visible(locator: Locator) -> bool:
    try:
        count = locator.count()
    except Exception:
        return False
    for idx in range(count):
        try:
            if locator.nth(idx).is_visible():
                return True
        except Exception:
            continue
    return False


def click_text_option(page: Page, candidates: list[str], timeout_ms: int = 10000) -> None:
    last_error: Exception | None = None
    for text in candidates:
        locator = page.get_by_text(text, exact=True)
        try:
            if click_first_visible(locator, timeout_ms=timeout_ms):
                return
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"未找到可点击的文本选项: {candidates}") from last_error


def wait_for_login(page: Page, home_url: str, timeout_ms: int = 300000) -> None:
    page.goto(home_url, wait_until="domcontentloaded", timeout=60000)
    if "signin" not in page.url:
        return
    print("检测到未登录金数据，请在弹出的浏览器里完成登录，脚本会自动继续。")
    try:
        page.wait_for_url(lambda url: "signin" not in url, timeout=timeout_ms)
        page.wait_for_load_state("domcontentloaded", timeout=60000)
    except Exception as exc:
        raise RuntimeError("等待金数据登录超时") from exc


def click_form_data_button(page: Page, form_title: str) -> None:
    page.wait_for_function(
        """
        (title) => {
          const normalize = (text) => (text || '').replace(/\\s+/g, ' ').trim();
          return [...document.querySelectorAll('*')].some((el) => normalize(el.textContent) === title || normalize(el.textContent).includes(title));
        }
        """,
        arg=form_title,
        timeout=30000,
    )
    result = page.evaluate(
        """
        ({ title }) => {
          const isVisible = (el) => {
            const style = window.getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style && style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
          };
          const normalize = (text) => (text || '').replace(/\\s+/g, ' ').trim();
          const all = [...document.querySelectorAll('*')].filter((el) => isVisible(el));
          const exactMatches = all.filter((el) => el.childElementCount === 0 && normalize(el.textContent) === title);
          const fuzzyMatches = all.filter((el) => el.childElementCount === 0 && normalize(el.textContent).includes(title));
          const titleElement = (exactMatches.length ? exactMatches : fuzzyMatches)[0];
          if (!titleElement) {
            return { ok: false, reason: 'title-not-found' };
          }
          let cur = titleElement;
          while (cur && cur !== document.body) {
            const entryLinks = [...cur.querySelectorAll('a[href*="/entries"]')].filter((el) => isVisible(el));
            if (entryLinks.length) {
              entryLinks[0].click();
              return { ok: true, mode: 'entries-link' };
            }
            const candidates = [...cur.querySelectorAll('button,a,[role="button"]')].filter((el) => isVisible(el) && el.innerText.trim() === '数据');
            if (candidates.length) {
              candidates[0].click();
              return { ok: true, mode: 'data-button' };
            }
            cur = cur.parentElement;
          }
          return { ok: false, reason: 'entries-link-not-found' };
        }
        """,
        {"title": form_title},
    )
    if not result.get("ok"):
        raise RuntimeError(f"无法在首页找到表单“{form_title}”的数据按钮: {result.get('reason')}")


def open_entries_page(page: Page, home_url: str, form_title: str | None, entries_url: str | None) -> None:
    if entries_url:
        page.goto(entries_url, wait_until="domcontentloaded", timeout=60000)
        try:
            page.wait_for_load_state("networkidle", timeout=5000)
        except Exception:
            pass
        page.locator("[data-testid='entry-grid__toolbar']").first.wait_for(timeout=30000)
        return
    wait_for_login(page, home_url)
    try:
        page.wait_for_load_state("networkidle", timeout=5000)
    except Exception:
        pass
    page.wait_for_timeout(3000)
    if not form_title:
        raise RuntimeError("未提供 form_title，无法从首页定位表单")
    click_form_data_button(page, form_title)
    try:
        page.wait_for_load_state("networkidle", timeout=5000)
    except Exception:
        pass
    page.locator("[data-testid='entry-grid__toolbar']").first.wait_for(timeout=30000)


def click_more_and_export(page: Page) -> bool:
    toolbar = page.locator("[data-testid='entry-grid__toolbar']").first
    toolbar.wait_for(timeout=30000)

    existing_download = page.locator(".grid-toolbar__export-excel a[data-role='download']")
    if has_visible(existing_download):
        return False
    existing_waiting = page.locator(".grid-toolbar__export-excel .export-excel-waiting")
    if existing_waiting.count():
        try:
            existing_waiting.first.wait_for(state="visible", timeout=1000)
            return False
        except Exception:
            pass

    direct_export = toolbar.get_by_role("button", name="导出", exact=True)
    if click_first_visible(direct_export, timeout_ms=5000):
        return True

    direct_export_text = toolbar.get_by_text("导出", exact=True)
    if click_first_visible(direct_export_text, timeout_ms=5000):
        return True

    more_button = toolbar.locator(
        "xpath=.//*[@id='open-search-entries-btn']"
        "/ancestor::div[contains(@class,'QueryAndShareAction_share-entries__yBqHC')]"
        "/following-sibling::*[1]//div[contains(@class,'ant-dropdown-trigger')]/button"
    ).first
    if not more_button.count():
        raise RuntimeError("未找到数据页工具栏中的三点菜单按钮")

    clicked = click_first_visible(more_button, timeout_ms=5000)
    if not clicked:
        raise RuntimeError("找到三点菜单按钮但点击失败")
    page.wait_for_function(
        """
        () => [...document.querySelectorAll('li[role="menuitem"]')].some((el) => {
          const style = window.getComputedStyle(el);
          const rect = el.getBoundingClientRect();
          return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
        })
        """,
        timeout=5000,
    )
    page.wait_for_function(
        """
        () => {
          const el = document.querySelector("li.ant-dropdown-menu-item[data-menu-id*='exportEntry']");
          if (!el) return false;
          const style = window.getComputedStyle(el);
          const rect = el.getBoundingClientRect();
          return style.visibility !== 'hidden'
            && style.display !== 'none'
            && rect.width > 0
            && rect.height > 0
            && el.getAttribute('aria-disabled') !== 'true'
            && !el.classList.contains('ant-dropdown-menu-item-disabled');
        }
        """,
        timeout=10000,
    )

    export_menu_item = page.locator("li.ant-dropdown-menu-item[data-menu-id*='exportEntry']")
    if click_first_visible(export_menu_item, timeout_ms=5000):
        return True
    fallback_menu_item = page.locator("li.ant-dropdown-menu-item").nth(1)
    if click_first_visible(fallback_menu_item, timeout_ms=5000):
        return True
    raise RuntimeError("已打开三点菜单，但未找到“导出数据”菜单项")


def pick_download_format(page: Page, download_format: str) -> None:
    format_label = (
        "label[for='export_job_sheet_format_excel']"
        if download_format == "xlsx"
        else "label[for='export_job_sheet_format_csv']"
    )
    page.locator(format_label).first.wait_for(state="visible", timeout=15000)
    if not click_first_visible(page.locator(format_label), timeout_ms=5000):
        raise RuntimeError(f"未找到可点击的导出格式选项: {download_format}")

    confirm_button = page.locator("a.submit.gd-btn.gd-btn-primary.second-step-el")
    if not click_first_visible(confirm_button, timeout_ms=5000):
        raise RuntimeError("未找到导出配置弹窗中的“确定”按钮")

    privacy_confirm = page.locator("#export_privacy_confirm_modal a.confirm")
    try:
        privacy_confirm.first.wait_for(state="visible", timeout=10000)
    except Exception:
        return
    if not click_first_visible(privacy_confirm, timeout_ms=5000):
        raise RuntimeError("未找到导出提示弹窗中的“继续导出”按钮")


def download_export_file(page: Page, run_dir: Path, download_format: str) -> Path:
    extension = ".xlsx" if download_format == "xlsx" else ".csv"
    page.wait_for_function(
        """
        () => {
          const isVisible = (el) => {
            const style = window.getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
          };
          return [...document.querySelectorAll(".grid-toolbar__export-excel a[data-role='download']")].some((el) => isVisible(el));
        }
        """,
        timeout=240000,
    )
    download_button = page.locator(".grid-toolbar__export-excel a[data-role='download']")
    with page.expect_download(timeout=120000) as download_info:
        if not click_first_visible(download_button, timeout_ms=5000):
            raise RuntimeError("下载链接已生成，但未找到可点击的“下载”按钮")
    download = download_info.value
    target = run_dir / f"jinshuju_export{extension}"
    download.save_as(str(target))
    return target


def export_jinshuju_file(
    context: BrowserContext,
    run_dir: Path,
    form_title: str | None,
    entries_url: str | None,
    home_url: str,
    download_format: str,
) -> Path:
    page = context.pages[0] if context.pages else context.new_page()
    try:
        open_entries_page(page, home_url=home_url, form_title=form_title, entries_url=entries_url)
        needs_format_picker = click_more_and_export(page)
        if needs_format_picker:
            pick_download_format(page, download_format)
        export_path = download_export_file(page, run_dir, download_format)
        return export_path
    except Exception:
        save_debug_artifacts(page, run_dir, "jinshuju_debug")
        raise


def launch_context_with_fallback(playwright_obj: Any, profile_dir: Path, headless: bool) -> tuple[BrowserContext, Path]:
    launch_kwargs = {
        "headless": headless,
        "accept_downloads": True,
        "viewport": {"width": 1440, "height": 960},
    }
    try:
        context = playwright_obj.chromium.launch_persistent_context(str(profile_dir), **launch_kwargs)
        return context, profile_dir
    except Error as exc:
        fallback_dir = Path(mkdtemp(prefix="jinshuju_profile_", dir=str(profile_dir.parent)))
        context = playwright_obj.chromium.launch_persistent_context(str(fallback_dir), **launch_kwargs)
        print(f"默认浏览器登录目录被占用，已切换到新的临时目录: {fallback_dir}")
        return context, fallback_dir


def write_outputs(
    run_dir: Path,
    matched: list[dict[str, str]],
    qualified_not_registered: list[QualificationRecord],
    registered_not_qualified: list[FormRecord],
    qualification_duplicates: list[DuplicateRow],
    form_duplicates: list[DuplicateRow],
    summary: dict[str, Any],
) -> None:
    write_csv(run_dir / "matched.csv", matched)
    write_csv(
        run_dir / "qualified_not_registered.csv",
        [qualification_to_csv_row(item) for item in qualified_not_registered],
    )
    write_csv(
        run_dir / "registered_not_qualified.csv",
        [form_to_csv_row(item) for item in registered_not_qualified],
    )
    write_csv(
        run_dir / "qualification_duplicates.csv",
        [duplicate_to_csv_row(item) for item in qualification_duplicates],
    )
    write_csv(
        run_dir / "form_duplicates.csv",
        [duplicate_to_csv_row(item) for item in form_duplicates],
    )
    (run_dir / "summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def print_summary(summary: dict[str, Any]) -> None:
    print(f"输出目录: {summary['output_dir']}")
    print(f"导出文件: {summary['export_file']}")
    print(f"资格名单总人数: {summary['qualification_total']}")
    print(f"金数据有效报名人数: {summary['form_total']}")
    print(f"匹配成功人数: {summary['matched_total']}")
    print(f"有资格但没报名: {summary['qualified_not_registered_total']}")
    print(f"没资格但报名了: {summary['registered_not_qualified_total']}")
    print(f"资格名单重复记录: {summary['qualification_duplicates_total']}")
    print(f"金数据重复记录: {summary['form_duplicates_total']}")


def main() -> int:
    args = parse_args()
    created_after = parse_user_datetime(args.created_after) if args.created_after else None
    run_dir = ensure_run_dir(args.output_root)
    qualifications = load_qualification_records(args.qualification_file)

    args.profile_dir.mkdir(parents=True, exist_ok=True)
    with sync_playwright() as p:
        context, active_profile_dir = launch_context_with_fallback(p, args.profile_dir, args.headless)
        try:
            export_path = export_jinshuju_file(
                context=context,
                run_dir=run_dir,
                form_title=args.form_title,
                entries_url=args.entries_url,
                home_url=args.home_url,
                download_format=args.download_format,
            )
        finally:
            context.close()

    form_records = load_form_records_from_export(
        export_path,
        name_label=args.name_field_label,
        college_label=args.college_field_label,
        qq_label=args.qq_field_label,
        created_after=created_after,
    )
    deduped_qualifications, qualification_duplicates = dedupe_qualifications(qualifications)
    deduped_form_records, form_duplicates = dedupe_form_records(form_records)
    matched, qualified_not_registered, registered_not_qualified = compare_records(
        deduped_qualifications,
        deduped_form_records,
    )

    summary = {
        "generated_at": datetime.now().astimezone().isoformat(),
        "output_dir": str(run_dir),
        "export_file": str(export_path),
        "form_title": args.form_title or "",
        "entries_url": args.entries_url or "",
        "qualification_file": str(args.qualification_file),
        "name_field_label": args.name_field_label,
        "college_field_label": args.college_field_label,
        "qq_field_label": args.qq_field_label,
        "created_after": args.created_after or "",
        "download_format": args.download_format,
        "profile_dir": str(active_profile_dir),
        "qualification_total": len(deduped_qualifications),
        "form_total": len(deduped_form_records),
        "matched_total": len(matched),
        "qualified_not_registered_total": len(qualified_not_registered),
        "registered_not_qualified_total": len(registered_not_qualified),
        "qualification_duplicates_total": len(qualification_duplicates),
        "form_duplicates_total": len(form_duplicates),
    }
    write_outputs(
        run_dir,
        matched=matched,
        qualified_not_registered=qualified_not_registered,
        registered_not_qualified=registered_not_qualified,
        qualification_duplicates=qualification_duplicates,
        form_duplicates=form_duplicates,
        summary=summary,
    )
    print_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
